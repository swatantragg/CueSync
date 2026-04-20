import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BORDER = Border(*(Side(style="thin"),) * 4)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def new_wb(title: str) -> tuple[Workbook, "Worksheet"]:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    return wb, ws


def save_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def set_row(ws, row_idx: int, values: list, bold=False, fill=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.border = BORDER
        c.alignment = CENTER
        if bold:
            c.font = HEADER_FONT
        if fill:
            c.fill = HEADER_FILL


def fmt_dur(sec: int | None) -> str:
    if not sec:
        return "00:00:00"
    h, rem = divmod(int(sec), 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"
