from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.cue_rules import iprs_role_code, role_key
from app.services.exporters._common import fmt_dur, save_bytes

# ── Borders ───────────────────────────────────────────────────────────────────
_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_NONE   = Side(style=None)

# ── Fills ─────────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

# ── Alignments ────────────────────────────────────────────────────────────────
_TOP_LEFT     = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
_TOP_LEFT_NW  = Alignment(horizontal="left",   vertical="top",    wrap_text=False)
_TOP_CENTER   = Alignment(horizontal="center", vertical="top",    wrap_text=True)
_TOP_CENTER_NW = Alignment(horizontal="center", vertical="top",   wrap_text=False)
_MID_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=False)

# ── Fonts ─────────────────────────────────────────────────────────────────────
_LBL    = Font(name="Calibri", bold=True,  size=10)   # label (col A, G)
_LBL11  = Font(name="Calibri", bold=True,  size=11)   # row-14 section labels
_VAL_B  = Font(name="Calibri", bold=True,  size=10)   # metadata value left side (B-F) — bold per reference
_VAL    = Font(name="Calibri", bold=False, size=10)   # metadata value right side (I-M)
_DAT    = Font(name="Calibri", bold=False, size=10)   # data row cols A-C, G-L
_DAT11  = Font(name="Calibri", bold=False, size=11)   # data row cols D, E, F, M — size 11 per reference

# ── Column widths — exact from reference file ──────────────────────────────────
_COL_WIDTHS = {
    "A": 43.71, "B": 8.43,  "C": 10.71, "D": 14.43, "E": 9.71,
    "F": 12.29, "G": 7.43,  "H": 24.29, "I": 11.0,  "J": 9.14,
    "K": 13.43, "L": 15.86, "M": 16.29,
}

# ── IPRS header rich text (36pt title + 11pt contact) ─────────────────────────
_IPRS_TITLE_TEXT = "The Indian Performing Right Society Limited"
_IPRS_CONTACT_TEXT = (
    "\nRegd. Office: 208, Golden Chambers, 2nd Floor, New Andheri Link Road, "
    "Andheri (W), Mumbai - 400 053.\n"
    "Tel: IPRS Call Center Number: 8097539960 & Office No. (091 22) 2673 3748 / 49 / 50 / 6616. "
    "Email: documentation@iprs.org, support.documentation@iprs.org. Visit us at: www.iprs.org\n"
    "CIN: U92140MH1969GAP014359"
)
_IPRS_RICH_HEADER = CellRichText(
    TextBlock(InlineFont(b=True, rFont="Arial Narrow", sz=36), _IPRS_TITLE_TEXT),
    TextBlock(InlineFont(b=True, rFont="Arial Narrow", sz=11), _IPRS_CONTACT_TEXT),
)

# ── Codes legend ──────────────────────────────────────────────────────────────
_CODES_TABLE = [
    ["Codes:", "BI -",  "Background Instrumental", "OI -", "Opening Instrumental",  "ROLE (C/A/E)"],
    ["",       "BV-",   "BackGround Vocal",         "OV -", "Opening Vocal",          "Composer"],
    ["",       "FI -",  "Feature Instrumental",     "CI -", "Closing Instrumental",   "Author/Lyricist"],
    ["",       "FV-",   "Feature Vocal",            "CV -", "Closing Vocal",          "Publisher"],
]


def _b(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN) -> Border:
    return Border(top=top, bottom=bottom, left=left, right=right)


def _block_border(row, col, r0, r1, c0, c1) -> Border:
    return Border(
        top    = _MEDIUM if row == r0 else _THIN,
        bottom = _MEDIUM if row == r1 else _THIN,
        left   = _MEDIUM if col == c0 else _THIN,
        right  = _MEDIUM if col == c1 else _THIN,
    )


def _apply_block_borders(ws, r0: int, r1: int, c0: int, c1: int) -> None:
    for row in range(r0, r1 + 1):
        for col in range(c0, c1 + 1):
            ws.cell(row=row, column=col).border = _block_border(row, col, r0, r1, c0, c1)


def _iprs_society(s: str | None) -> str:
    v = (s or "IPRS").strip().upper()
    if v in ("NS", "NON SOCIETY", "NON-SOCIETY") or v.startswith("NON"):
        return "NS"
    return s or "IPRS"


def _sorted_contribs(contribs: list) -> list:
    _order = {"Composer": 0, "CA": 0, "Author": 1, "Publisher": 2}
    return sorted(contribs, key=lambda c: _order.get(role_key(c.role), 3))


# Per-column (data rows) font and alignment — derived from reference
# col: (font, alignment)
_DATA_STYLE = {
    1:  (_DAT,   _TOP_LEFT),       # song title: left, wrap
    2:  (_DAT,   _TOP_CENTER),     # characteristics: center, wrap
    3:  (_DAT,   _TOP_CENTER),     # no of usage: center, wrap
    4:  (_DAT11, _TOP_CENTER),     # song code: center, wrap, size 11
    5:  (_DAT11, _TOP_CENTER),     # ISRC: center, wrap, size 11
    6:  (_DAT11, _TOP_CENTER_NW),  # duration: center, NO wrap, size 11
    7:  (_DAT,   _TOP_CENTER_NW),  # role: center, NO wrap
    8:  (_DAT,   _TOP_LEFT_NW),    # names: left, NO wrap
    9:  (_DAT,   _MID_CENTER),     # society: center/center, NO wrap
    10: (_DAT,   _TOP_CENTER),     # share: center, wrap
    11: (_DAT,   _TOP_CENTER_NW),  # IPI: center, NO wrap
    12: (_DAT,   _TOP_LEFT_NW),    # singer: left, NO wrap (NOT merged)
    13: (_DAT11, _TOP_LEFT_NW),    # validation: left, NO wrap, size 11 (merged)
}


def _fill_iprs_sheet(ws, project, episode, total_ep_count: int | None = None) -> None:
    # ── Row 1: IPRS header ────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws.cell(row=1, column=1)
    c.value = _IPRS_RICH_HEADER
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 92.25

    # ── Row 2: Cue sheet type title ───────────────────────────────────────────
    ws.merge_cells("A2:M2")
    c = ws.cell(row=2, column=1, value="TV/WEB SERIES CUE SHEET")
    c.font = Font(name="Calibri", bold=True, size=28)
    c.alignment = _TOP_CENTER
    ws.row_dimensions[2].height = 36.75

    # ── Metadata values ───────────────────────────────────────────────────────
    serial_title = episode.cue_serial_title or project.title or ""
    channel      = episode.cue_channel or project.channel_name or ""
    serial_type  = episode.cue_serial_type or ""
    language     = episode.cue_language or project.language or ""
    director     = episode.cue_director or project.director or ""
    genre        = episode.cue_genre or project.genre or ""
    prod_company = episode.cue_production_company or project.production_company or ""
    actors       = episode.cue_actors or project.actors or ""
    prod_year    = episode.cue_production_year or project.production_year or ""
    producer     = episode.cue_producer or project.producer or ""
    bg_composer  = episode.cue_bg_music_composer or project.bg_music_composer or ""
    submitted_by = episode.cue_submitted_by or project.submitted_by or ""
    air_date     = str(episode.air_date) if episode.air_date else ""

    def _meta(row: int, ll: str, lv, rl: str, rv) -> None:
        ws.merge_cells(f"B{row}:F{row}")
        ws.merge_cells(f"G{row}:H{row}")
        ws.merge_cells(f"I{row}:M{row}")
        # Col A — label, bold, left
        c = ws.cell(row=row, column=1, value=ll)
        c.font = _LBL; c.alignment = _TOP_LEFT
        # Col B (B:F merged) — value, bold, CENTER per reference
        c = ws.cell(row=row, column=2, value=lv)
        c.font = _VAL_B; c.alignment = _TOP_CENTER
        # Col G (G:H merged) — right-side label, left
        c = ws.cell(row=row, column=7, value=rl)
        c.font = _LBL; c.alignment = _TOP_LEFT
        # Col I (I:M merged) — right-side value, NOT bold, CENTER per reference
        c = ws.cell(row=row, column=9, value=rv)
        c.font = _VAL; c.alignment = _TOP_CENTER

    _meta(3,  "SERIAL TITLE", serial_title,
              "CHANNEL NAME", channel)
    _meta(4,  "SERIAL TYPE [EG: DOCUMENTRY, SOAPS, WEB SERIES]", serial_type,
              "DIRECTOR", director)
    _meta(5,  "GENRE / CATEGORY", genre,
              "BANNER / PRODUCTION COMPANY", prod_company)
    _meta(6,  "LANGUAGE", language,
              "PRINCIPAL ACTORS / ACTRESS", actors)
    _meta(7,  "PRODUCTION NUMBER", "",
              "TOTAL EPISODE DURATION", fmt_dur(episode.total_duration_sec))
    _meta(8,  "DATE OF EPISODE 1ST PERFORMED / AIRED", air_date,
              "TOTAL MUSICAL DURATION", fmt_dur(episode.musical_duration_sec))
    _meta(9,  "PRODUCTION YEAR", str(prod_year) if prod_year else "",
              "BACKGROUND MUSIC COMPOSER", bg_composer)
    _meta(10, "PRODUCER", producer,
              "Submitted By (Name of C/A/E)", submitted_by)

    # Row 11: Episode No / Title — same column structure as _meta
    ws.merge_cells("B11:F11")
    ws.merge_cells("G11:H11")
    ws.merge_cells("I11:M11")
    for col, val, fnt, aln in (
        (1, "EPISODE NO.",          _LBL,   _TOP_LEFT),
        (2, episode.episode_number, _VAL_B, _TOP_CENTER),
        (7, "EPISODE TITLE",        _LBL,   _TOP_LEFT),
        (9, episode.title or "",    _VAL,   _TOP_CENTER),
    ):
        c = ws.cell(row=11, column=col, value=val)
        c.font = fnt; c.alignment = aln

    # Row 12: Total episodes — same column structure (NOT merged B:M)
    ws.merge_cells("B12:F12")
    ws.merge_cells("G12:H12")
    ws.merge_cells("I12:M12")
    total_eps_val = total_ep_count or project.total_episodes or ""
    for col, val, fnt, aln in (
        (1, "TOTAL NO. OF EPISODE", _LBL,   _TOP_LEFT),
        (2, total_eps_val,          _VAL_B, _TOP_CENTER),
        (7, "",                     _LBL,   _TOP_LEFT),
        (9, "",                     _VAL,   _TOP_CENTER),
    ):
        c = ws.cell(row=12, column=col, value=val)
        c.font = fnt; c.alignment = aln

    # Medium outside + thin inside for whole metadata block (rows 3-12)
    _apply_block_borders(ws, 3, 12, 1, 13)

    # ── Row 13: blank separator ───────────────────────────────────────────────
    ws.merge_cells("A13:M13")
    ws.row_dimensions[13].height = 15.75

    # ── Row 14: Section labels ────────────────────────────────────────────────
    ws.merge_cells("A14:F14")
    ws.merge_cells("G14:K14")
    ws.merge_cells("L14:M14")
    ws.row_dimensions[14].height = 15.75
    c = ws.cell(row=14, column=1, value="WORK DETAILS")
    c.font = _LBL11; c.alignment = _TOP_CENTER
    c = ws.cell(row=14, column=7, value="COMPOSER / AUTHOR / PUBLISHER / SINGER DETAILS")
    c.font = _LBL11; c.alignment = _TOP_CENTER
    _apply_block_borders(ws, 14, 14, 1, 6)
    _apply_block_borders(ws, 14, 14, 7, 11)
    _apply_block_borders(ws, 14, 14, 12, 13)

    # ── Row 15: Column headers ────────────────────────────────────────────────
    _HEADERS = [
        "SONG TITLE / TRACK NAME",
        "CHARACTERISTICS",
        "NO. OF USAGE",
        "Internal No/Song Code",
        "ISRC",
        "DURATION\n(HH:MM:SS)",
        "ROLE\n(C / A / E)",
        "NAMES OF COMPOSER / AUTHOR / PUBLISHER",
        "SOCIETY",
        "SHARE",
        "IPI NO.",
        "SINGER",
        "VALIDATION LINK (IF SONG CODE IS NOT AVAILABLE)",
    ]
    ws.row_dimensions[15].height = 38.25
    _MED_ALL = _b(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)
    for col, hdr in enumerate(_HEADERS, 1):
        c = ws.cell(row=15, column=col, value=hdr)
        c.font = _LBL
        c.fill = _HEADER_FILL
        c.border = _MED_ALL
        # Col 13 (validation): left-aligned per reference; others center
        c.alignment = _TOP_LEFT if col == 13 else _TOP_CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    cues_ordered = sorted(episode.cues, key=lambda c: (c.order_index, c.id))
    r = 16
    for cue in cues_ordered:
        contribs = _sorted_contribs(list(cue.contributors or []))
        n_rows   = max(len(contribs), 1)

        # Merge cols A-F and M (validation) across contributor rows.
        # Col L (SINGER) is NOT merged per reference — singer only on first row.
        if n_rows > 1:
            for col in (1, 2, 3, 4, 5, 6, 13):
                ws.merge_cells(
                    start_row=r, start_column=col,
                    end_row=r + n_rows - 1, end_column=col,
                )

        for i, contrib in enumerate(contribs or [None]):
            first = i == 0
            share_str = (
                f"{float(contrib.share_percent or 0):.2f}%"
                if contrib else None
            )
            row_vals = [
                cue.song_title                                          if first else None,  # 1 A
                (cue.usage_type.value.upper() if cue.usage_type else "") if first else None, # 2 B
                str(cue.usage_count or 1)                               if first else None,  # 3 C
                cue.song_code                                           if first else None,  # 4 D
                cue.isrc                                                if first else None,  # 5 E
                fmt_dur(cue.duration_sec)                               if first else None,  # 6 F
                iprs_role_code(contrib.role)                            if contrib else None, # 7 G
                contrib.name                                            if contrib else None, # 8 H
                _iprs_society(contrib.society)                          if contrib else None, # 9 I
                share_str,                                                                    # 10 J
                str(contrib.ipi_number or contrib.cae_number or "")    if contrib else None, # 11 K
                cue.singer                                              if first else None,   # 12 L (not merged)
                cue.validation_link                                     if first else None,   # 13 M (merged)
            ]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                fnt, aln = _DATA_STYLE[col]
                c.font = fnt; c.alignment = aln
            ws.row_dimensions[r].height = 14.25
            r += 1

        # Block borders for this song
        _apply_block_borders(ws, r - n_rows, r - 1, 1, 13)

        # Col G (role) left border = None per reference
        for data_row in range(r - n_rows, r):
            cell = ws.cell(row=data_row, column=7)
            b = cell.border
            cell.border = Border(top=b.top, bottom=b.bottom, left=_NONE, right=b.right)

    # ── Footer ────────────────────────────────────────────────────────────────
    foot = r + 1  # one blank row then codes start
    for i, row_data in enumerate(_CODES_TABLE):
        curr = foot + i
        ws.row_dimensions[curr].height = 15.75
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=curr, column=col, value=val)
            c.font = _LBL if (col == 1 and i == 0) else Font(name="Calibri", size=10)
            # Wrap text ON for codes (wide text wraps inside narrow column C)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _apply_block_borders(ws, foot, foot + len(_CODES_TABLE) - 1, 1, 6)

    note_row = foot + len(_CODES_TABLE) + 2
    c = ws.cell(row=note_row, column=1,
                value="Note: Cue Sheets should be submitted for each Language Separately.")
    c.font = Font(name="Calibri", bold=True, size=10)
    # NO wrap_text per reference
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    prep_row = note_row + 2
    c = ws.cell(row=prep_row, column=1,
                value="Cue Sheet Prepared By : Samraj Music Rights Managment Pvt. Ltd")
    c.font = Font(name="Calibri", bold=True, size=20)
    # NO wrap_text per reference
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Column widths (exact from reference) ──────────────────────────────────
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def build_iprs(project, episode, total_ep_count: int | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = str(episode.episode_number)[:31]
    _fill_iprs_sheet(ws, project, episode, total_ep_count=total_ep_count)
    return save_bytes(wb)


def build_iprs_bulk(project, episodes, total_ep_count: int | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for ep in sorted(episodes, key=lambda e: e.episode_number):
        ws = wb.create_sheet(title=str(ep.episode_number)[:31])
        _fill_iprs_sheet(ws, project, ep, total_ep_count=total_ep_count)
    return save_bytes(wb)
