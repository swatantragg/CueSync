from datetime import datetime as _dt, time as _time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.cue_rules import computed_share, role_key, sorted_cues, usage_code
from app.services.exporters._common import save_bytes

# ── Constants ─────────────────────────────────────────────────────────────────
_TNR = "Times New Roman"
_ANW = "Arial Narrow"
_GRAY = PatternFill("solid", fgColor="D7D7D7")

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_NONE   = Side(style=None)

# exact alignments extracted from reference xlsx
_TOP_GEN    = Alignment(vertical="top")                                         # h=None (General)
_TOP_LEFT   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
_TOP_LEFT_NW = Alignment(horizontal="left",  vertical="top", wrap_text=False)
_TOP_CTR    = Alignment(horizontal="center", vertical="top", wrap_text=False)   # center, no wrap
_TOP_CTR_W  = Alignment(horizontal="center", vertical="top", wrap_text=True)   # center + wrap

# ── Fonts ─────────────────────────────────────────────────────────────────────
_F_TNR_15B = Font(name=_TNR, bold=True,  size=15)
_F_TNR_11B = Font(name=_TNR, bold=True,  size=11)
_F_TNR_11  = Font(name=_TNR, bold=False, size=11)
_F_TNR_10B = Font(name=_TNR, bold=True,  size=10)
_F_ANW_10B = Font(name=_ANW, bold=True,  size=10)
_F_ANW_10  = Font(name=_ANW, bold=False, size=10)
_F_ANW_11  = Font(name=_ANW, bold=False, size=11)
_F_ANW_15B = Font(name=_ANW, bold=True,  size=15)

# ── Borders ───────────────────────────────────────────────────────────────────
_MED_ALL  = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)
_THIN_ALL = Border(top=_THIN,   bottom=_THIN,   left=_THIN,   right=_THIN)


def _b(top=_NONE, bottom=_NONE, left=_NONE, right=_NONE) -> Border:
    return Border(top=top, bottom=bottom, left=left, right=right)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _secs_to_time(sec) -> _time | None:
    if not sec:
        return None
    h, rem = divmod(int(sec), 3600)
    m, s = divmod(rem, 60)
    return _time(h % 24, m, s)


def _parse_air_date(val) -> _dt | None:
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return _dt.strptime(str(val), fmt)
        except ValueError:
            continue
    return None


def _prs_usage(cue) -> str:
    code = usage_code(cue.usage_type, cue.song_title)
    return "B" if code == "BI" else "F"


def _prs_role(role) -> str:
    k = role_key(role)
    return {"Composer": "Composer", "Author": "Author",
            "Publisher": "Publisher", "CA": "Composer/Author"}.get(k, k or "Composer")


def _prs_society(s) -> str:
    v = (s or "IPRS").strip().upper()
    if v in ("", "IPRS"):
        return "PRS"
    if v in ("NS", "NON SOCIETY", "NON-SOCIETY") or v.startswith("NON"):
        return "NS"
    return s or "PRS"


def _prs_share(val) -> str:
    return f"{float(val or 0):.0f}%"


def _sort_contribs(contribs: list) -> list:
    _ord = {"Composer": 0, "CA": 0, "Author": 1, "Publisher": 2}
    return sorted(contribs, key=lambda c: _ord.get(role_key(getattr(c, "role", None)), 3))


# ── Header builder ────────────────────────────────────────────────────────────
def _write_header(ws, project, episode) -> None:
    serial_title = episode.cue_serial_title or project.title or ""
    channel      = episode.cue_channel      or project.channel_name or ""
    director     = episode.cue_director     or project.director or ""
    prod_company = episode.cue_production_company or project.production_company or ""
    actors       = episode.cue_actors       or project.actors or ""
    prod_year    = episode.cue_production_year or project.production_year or ""
    country      = episode.cue_country      or project.country or "India"
    ep_number    = episode.episode_number
    ep_title     = episode.title or ""
    air_date     = _parse_air_date(episode.air_date)
    total_dur    = _secs_to_time(episode.total_duration_sec)
    music_dur    = _secs_to_time(episode.musical_duration_sec)

    # ── Row 1 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "LICENSOR APPROVED MUSIC CUE SHEET"
    c.font = _F_TNR_15B
    c.alignment = _TOP_GEN
    ws.row_dimensions[1].height = 18

    # ── Row 2 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A2:L2")
    c = ws["A2"]
    c.value = "Important Note"
    c.font = _F_TNR_11B
    c.alignment = _TOP_GEN

    # ── Row 3 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A3:L3")
    c = ws["A3"]
    c.value = (
        "The correct completion of this cue sheet is a strict condition of the granting of "
        "the broadcast licences by the various copyright owners and agencies. The shaded "
        "sections indicate essential items of information which must be supplied in all cases. "
        "The unshaded sections signify information which should be supplied if it is available."
    )
    c.font = _F_TNR_11
    c.alignment = _TOP_LEFT
    ws.row_dimensions[3].height = 15

    # ── Row 4 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = (
        "Please turn to the reverse side of this form for explanatory notes, a key to the "
        "standard codes, and helpful telephone contacts should you require any further assistance."
    )
    c.font = _F_TNR_11
    c.alignment = _TOP_LEFT
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 15

    # ── Row 6 ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A6:B6")
    c = ws["A6"]
    c.value = "Production Details"
    c.font = _F_TNR_10B
    c.alignment = _TOP_LEFT
    ws.row_dimensions[6].height = 15

    # ── Rows 7-11: Production metadata ───────────────────────────────────────
    for r in range(7, 11):
        ws.row_dimensions[r].height = 54.95

    def _lbl(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = _F_TNR_10B
        c.fill = _GRAY
        c.border = _MED_ALL
        c.alignment = _TOP_LEFT

    def _val(row, col, value, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = _F_ANW_10
        c.border = _MED_ALL
        c.alignment = _TOP_LEFT
        if number_format:
            c.number_format = number_format
        return c

    def _apply_med_all_cols(row, col_start, col_end):
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).border = _MED_ALL

    # Row 7
    ws.merge_cells("A7:B7");  _lbl(7, 1, "Film/Series/Item Title")
    ws.merge_cells("C7:D7");  _val(7, 3, serial_title); _apply_med_all_cols(7, 3, 4)
    _lbl(7, 5, "Production Co.")
    _val(7, 6, prod_company)
    ws.merge_cells("G7:H7");  _lbl(7, 7, "Country of Origin"); _apply_med_all_cols(7, 7, 8)
    _val(7, 9, country)
    ws.merge_cells("J7:K7");  _lbl(7, 10, "Trailer /Promo/ Full Programme (T/P/F)"); _apply_med_all_cols(7, 10, 11)
    c = ws.cell(row=7, column=12, value="F")
    c.font = _F_ANW_10B; c.border = _MED_ALL; c.alignment = _TOP_LEFT

    # Row 8
    ws.merge_cells("A8:B8");  _lbl(8, 1, "Episode Title")
    ws.merge_cells("C8:D8");  _val(8, 3, ep_title); _apply_med_all_cols(8, 3, 4)
    _lbl(8, 5, "Production No.")
    _val(8, 6, None)
    ws.merge_cells("G8:H8");  _lbl(8, 7, "Production Year"); _apply_med_all_cols(8, 7, 8)
    _val(8, 9, prod_year)
    ws.merge_cells("J8:K8");  _lbl(8, 10, "Tx Time"); _apply_med_all_cols(8, 10, 11)
    _val(8, 12, None)

    # Row 9
    ws.merge_cells("A9:B9");  _lbl(9, 1, "Episode No.")
    ws.merge_cells("C9:D9");  _val(9, 3, ep_number); _apply_med_all_cols(9, 3, 4)
    _lbl(9, 5, "Director")
    _val(9, 6, director)
    ws.merge_cells("G9:H9");  _lbl(9, 7, "First Tx Date"); _apply_med_all_cols(9, 7, 8)
    _val(9, 9, air_date, number_format="DD MMMM YYYY")
    ws.merge_cells("J9:K9");  _lbl(9, 10, "Film Duration"); _apply_med_all_cols(9, 10, 11)
    _val(9, 12, total_dur, number_format="hh:mm:ss")

    # Row 10
    ws.merge_cells("A10:B10"); _lbl(10, 1, "Film/Item No.")
    ws.merge_cells("C10:D10"); _val(10, 3, None); _apply_med_all_cols(10, 3, 4)
    _lbl(10, 5, "Principal Actors")
    _val(10, 6, actors)
    ws.merge_cells("G10:H10"); _lbl(10, 7, "Channel"); _apply_med_all_cols(10, 7, 8)
    _val(10, 9, channel)
    ws.merge_cells("J10:K10"); _lbl(10, 10, "Music Duration"); _apply_med_all_cols(10, 10, 11)
    _val(10, 12, music_dur, number_format="hh:mm:ss")

    # Row 11
    ws.row_dimensions[11].height = 15.75
    ws.merge_cells("A11:B11"); _lbl(11, 1, "Alternative Title(s)")
    ws.merge_cells("C11:D11"); _apply_med_all_cols(11, 3, 4)
    # E11: gray fill + medium borders (E7:E11 all get gray per VBA)
    c = ws.cell(row=11, column=5); c.fill = _GRAY; c.border = _MED_ALL
    # F11: medium borders, no fill
    ws.cell(row=11, column=6).border = _MED_ALL
    ws.merge_cells("G11:L11"); _apply_med_all_cols(11, 7, 12)

    # Row 13
    ws.row_dimensions[13].height = 15.75
    ws.merge_cells("A13:B13")
    c = ws["A13"]
    c.value = "Music Details"
    c.font = _F_TNR_10B
    c.alignment = _TOP_LEFT

    # ── Row 14: Column headers ────────────────────────────────────────────────
    ws.row_dimensions[14].height = 14.1
    _HDR = [
        (1, None, "Seq"),
        (2, 3,    "Title"),
        (4, None, "Role"),
        (5, 6,    "Interested Parties"),
        (7, None, "CAE Number"),
        (8, None, "Share %"),
        (9, None, "Society"),
        (10, None, "Usage"),
        (11, None, "Duration"),
        (12, None, "Work Number"),
    ]
    for start_col, end_col, text in _HDR:
        if end_col:
            ws.merge_cells(
                start_row=14, start_column=start_col,
                end_row=14, end_column=end_col,
            )
            for col in range(start_col, end_col + 1):
                ws.cell(row=14, column=col).border = _MED_ALL
                ws.cell(row=14, column=col).fill = _GRAY
        c = ws.cell(row=14, column=start_col, value=text)
        c.font = _F_TNR_10B
        c.fill = _GRAY
        c.border = _MED_ALL
        c.alignment = _TOP_GEN           # h=None (General) — matches reference exactly
        if not end_col:
            ws.cell(row=14, column=start_col).border = _MED_ALL


# ── Data rows ─────────────────────────────────────────────────────────────────
def _write_data(ws, episode) -> int:
    """Write cue data rows starting at row 15. Returns the next empty row."""
    cues = sorted_cues(episode.cues)
    r = 15

    for idx, cue in enumerate(cues, 1):
        contribs = _sort_contribs(list(cue.contributors or []))
        n = max(len(contribs), 1)
        usage_val = _prs_usage(cue)
        dur_val   = _secs_to_time(cue.duration_sec)
        work_num  = cue.work_number or ""

        # Vertical merges for Seq (col A=1) and Usage (col J=10)
        if n > 1:
            ws.merge_cells(start_row=r, start_column=1,  end_row=r+n-1, end_column=1)
            ws.merge_cells(start_row=r, start_column=10, end_row=r+n-1, end_column=10)

        for i, contrib in enumerate(contribs or [None]):
            row_num  = r + i
            is_first = (i == 0)
            is_last  = (i == n - 1)
            ws.row_dimensions[row_num].height = 16.5

            # Top / bottom sides per row position
            top    = _THIN if is_first else _NONE
            bottom = _THIN if is_last  else _NONE
            # B/C/E/F top: medium ONLY on the absolute first data row (row 15, adjacent to
            # the medium-bordered header row 14). All subsequent cue first rows use thin,
            # which matches the reference file (B18/E18 top=thin, B15/E15 top=medium).
            if is_first:
                top_bc = _MEDIUM if r == 15 else _THIN
            else:
                top_bc = _NONE

            # ── Col A (1): Seq — merged vertically ───────────────────────────
            c = ws.cell(row=row_num, column=1)
            if is_first:
                c.value = idx
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR_W    # center + wrap — matches reference

            # ── Col B (2): Title — value on first row only ────────────────────
            c = ws.cell(row=row_num, column=2)
            c.value     = cue.song_title if is_first else None
            c.font      = _F_ANW_11
            c.border    = _b(top=top_bc, bottom=bottom, left=_THIN, right=_NONE)
            c.alignment = _TOP_GEN      # h=None — matches reference

            # ── Col C (3): empty — right edge of "Title" column ───────────────
            c = ws.cell(row=row_num, column=3)
            c.border    = _b(top=top_bc, bottom=bottom, left=_NONE, right=_THIN)

            # ── Col D (4): Role ───────────────────────────────────────────────
            c = ws.cell(row=row_num, column=4)
            c.value     = _prs_role(contrib.role) if contrib else ""
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR      # center no-wrap — matches reference

            # ── Col E (5): Interested Parties — left edge ─────────────────────
            c = ws.cell(row=row_num, column=5)
            c.value     = contrib.name if contrib else ""
            c.font      = _F_ANW_11
            c.border    = _b(top=top_bc, bottom=bottom, left=_THIN, right=_NONE)
            c.alignment = _TOP_GEN      # h=None — matches reference

            # ── Col F (6): empty — right edge of "Interested Parties" column ──
            c = ws.cell(row=row_num, column=6)
            c.border    = _b(top=top_bc, bottom=bottom, left=_NONE, right=_THIN)

            # ── Col G (7): CAE Number ─────────────────────────────────────────
            cae = (contrib.cae_number or contrib.ipi_number or "") if contrib else ""
            c = ws.cell(row=row_num, column=7)
            c.value     = cae
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR      # center no-wrap — matches reference

            # ── Col H (8): Share % ────────────────────────────────────────────
            share_disp = None
            if contrib:
                share_val  = computed_share(contrib, contribs)
                share_disp = _prs_share(share_val)
            c = ws.cell(row=row_num, column=8)
            c.value     = share_disp
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR_W    # center + wrap — matches reference

            # ── Col I (9): Society ────────────────────────────────────────────
            soc = _prs_society(contrib.society) if contrib else ""
            c = ws.cell(row=row_num, column=9)
            c.value     = soc
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR      # center no-wrap — matches reference

            # ── Col J (10): Usage — merged vertically ────────────────────────
            c = ws.cell(row=row_num, column=10)
            if is_first:
                c.value = usage_val
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR_W    # center + wrap — matches reference

            # ── Col K (11): Duration — first row only ────────────────────────
            c = ws.cell(row=row_num, column=11)
            c.value  = dur_val if is_first else None
            c.font   = _F_ANW_11
            c.border = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_CTR      # center no-wrap — matches reference
            if is_first and dur_val:
                c.number_format = "hh:mm:ss"

            # ── Col L (12): Work Number — first row only ──────────────────────
            c = ws.cell(row=row_num, column=12)
            c.value     = work_num if is_first else None
            c.font      = _F_ANW_11
            c.border    = _b(top=top, bottom=bottom, left=_THIN, right=_THIN)
            c.alignment = _TOP_GEN      # h=None — General

        r += n

    return r  # next empty row after all data


# ── Column widths ─────────────────────────────────────────────────────────────
_COL_WIDTHS = {
    "A": 5, "B": 24, "C": 21, "D": 15, "E": 17, "F": 34,
    "G": 12, "H": 7, "I": 13, "J": 5, "K": 11, "L": 12,
}


def _fill_prs_sheet(ws, project, episode) -> None:
    _write_header(ws, project, episode)
    next_row = _write_data(ws, episode)

    # ── Footer ────────────────────────────────────────────────────────────────
    c = ws.cell(row=next_row + 1, column=1,
                value="Cue Sheet Prepared By: Samraj Music Rights Management Pvt. Ltd.")
    c.font      = _F_ANW_15B
    c.alignment = _TOP_LEFT_NW

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Remove gridlines (matches reference — VBA: ActiveWindow.DisplayGridlines = False) ──
    ws.sheet_view.showGridLines = False


def build_prs(project, episode) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = str(episode.episode_number)[:31]
    _fill_prs_sheet(ws, project, episode)
    return save_bytes(wb)


def build_prs_bulk(project, episodes) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for ep in sorted(episodes, key=lambda e: e.episode_number):
        ws = wb.create_sheet(title=str(ep.episode_number)[:31])
        _fill_prs_sheet(ws, project, ep)
    return save_bytes(wb)
