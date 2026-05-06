from __future__ import annotations

from datetime import date, datetime
from datetime import time as dtime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.services.cue_rules import computed_share, role_key, sorted_cues, usage_code
from app.services.exporters._common import save_bytes

_THIN = Side(style="thin")


def _b_all() -> Border:
    return Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)


# Fonts from ASCAP reference workbook / VBA output.
_TITLE_FONT = Font(name="Arial", bold=True, size=14)
_META_FONT = Font(name="Arial", bold=True, size=11)
_HDR_FONT = Font(name="Calibri", bold=True, size=10)
_DATA_FONT = Font(name="Arial Narrow", bold=False, size=11)
_FOOT_FONT = Font(name="Calibri", bold=True, size=20)
_NOTE_FONT = Font(name="Calibri", bold=False, size=10)

_CENTER = Alignment(horizontal="center", wrap_text=False)
_WRAP_C = Alignment(horizontal="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", wrap_text=False)

# Widths as stored in the real ASCAP workbook after Excel ColumnWidth autofit.
_COL_WIDTHS = {
    "A": 6.7109375,
    "B": 48.7109375,
    "C": 6.7109375,
    "D": 10.7109375,
    "E": 47.7109375,
    "F": 58.7109375,
    "G": 18.7109375,
}

# Exact separator from reference: 100 spaces between multiple composers.
_SEP = " " * 100

_ASCAP_USE = {"BI": "BI", "BV": "BV", "FI": "VI", "FV": "VV", "OI": "MT", "CI": "ET"}

_USE_CODES = [
    "                *Use Codes: MT = Main Title      VI = Visual Instrumental    BV = Background Vocal",
    "                                          VV = Visual Vocal  ET = End Title                   BI = Background Instrumental",
    "                                          T = Theme",
]


def _fmt_date(air_date) -> str:
    if not air_date:
        return ""
    if isinstance(air_date, date):
        return air_date.strftime("%d %B %Y")
    try:
        return datetime.strptime(str(air_date), "%Y-%m-%d").strftime("%d %B %Y")
    except Exception:
        return str(air_date)


def _fmt_dur(duration_sec: int | None) -> str:
    sec = int(duration_sec or 0)
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _to_time(duration_sec: int | None) -> dtime:
    """Return datetime.time; reference stores timing as time object."""
    sec = int(duration_sec or 0)
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return dtime(min(h, 23), m, s)


def _society(raw: str | None) -> str:
    s = (raw or "").strip()
    return "Non Society" if s.upper() in ("NS", "(NS)") else (s or "IPRS")


def _work_id(raw):
    value = str(raw or "").strip()
    if value.isdigit() and not value.startswith("0"):
        return int(value)
    return value or None


def _fmt_contribs(contribs: list, cue, *, is_publisher: bool) -> str:
    use = usage_code(getattr(cue, "usage_type", None), getattr(cue, "song_title", None))
    parts = []
    for c in contribs:
        if is_publisher != (role_key(c.role) == "Publisher"):
            continue
        name = (c.name or "").strip()
        if not name:
            continue
        # ASCAP share = IPRS share x 2 (VBA: TEXT(J*2,"0.00%"), .00 stripped).
        share = computed_share(c, contribs, char_code=use) * 2
        parts.append(f"{name} ({_society(c.society)}) {share:.0f}%")
    return _SEP.join(parts)


def _data_row_height(row_data: list) -> float:
    title = str(row_data[1] or "")
    composers = str(row_data[4] or "")
    publishers = str(row_data[5] or "")
    has_wrapped_title = len(title) > 54
    has_wrapped_composers = _SEP in composers or "\n" in composers
    has_wrapped_publishers = len(publishers) > 62 or "\n" in publishers
    return 33.0 if (has_wrapped_title or has_wrapped_composers or has_wrapped_publishers) else 16.5


def _prepare_worksheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_format.defaultRowHeight = 15
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _fill_ascap_sheet(ws, project, episode) -> None:
    _prepare_worksheet(ws)

    serial_title = getattr(episode, "cue_serial_title", None) or project.title or ""
    ep_num       = episode.episode_number
    ep_title     = getattr(episode, "title", None) or ""
    duration     = _fmt_dur(getattr(episode, "total_duration_sec", None))
    contact      = (
        getattr(episode, "cue_submitted_by", None)
        or getattr(project, "submitted_by", None)
        or "NA"
    )

    # Row 1: title, merged A1:G1, no border.
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=f"{serial_title}(Hindi Serial)Cuesheet")
    c.font = _TITLE_FONT
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 23.25

    # Rows 3-7: plain metadata cells, no merges and no borders.
    ep_label = f"{ep_num} {ep_title}".strip() if ep_title else str(ep_num)
    for i, (lv, rv) in enumerate(zip(
        [
            f"Series/Film Title: {serial_title}",
            f"Episode Title/Number: {ep_label}",
            f"Estimated Airdate: {_fmt_date(getattr(episode, 'air_date', None))}",
            f"Program Length: {duration}",
            "Program Type: Soap",
        ],
        [
            "Company Name: NA",
            "Address: NA",
            "Phone: NA",
            f"Contact: {contact}",
            "Network Station: NA",
        ],
    ), 3):
        cl = ws.cell(row=i, column=1, value=lv)
        cl.font = _META_FONT; cl.alignment = _LEFT
        cr = ws.cell(row=i, column=5, value=rv)
        cr.font = _META_FONT; cr.alignment = _LEFT

    # Row 9: headers.
    for col, hdr in enumerate(
        ["Cue #", "Cue Title", "Use*", "Timing",
         "Composer(s) Affiliation / %", "Publisher(s) Affiliation / %", "Work ID (ASCAP)"],
        1,
    ):
        c = ws.cell(row=9, column=col, value=hdr)
        c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _b_all()

    # Data rows from row 10.
    r = 10
    for idx, cue in enumerate(sorted_cues(episode.cues), 1):
        contribs = list(cue.contributors or [])
        use = usage_code(cue.usage_type, cue.song_title)
        row_data = [
            idx,
            cue.song_title or "",
            _ASCAP_USE.get(use, use),
            _to_time(cue.duration_sec),                               # time object
            _fmt_contribs(contribs, cue, is_publisher=False),
            _fmt_contribs(contribs, cue, is_publisher=True),
            _work_id(cue.ascap_work_id),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _DATA_FONT
            c.border = _b_all()
            if col in (3, 4):
                c.number_format = "h:mm:ss"
            if col == 4:
                c.alignment = _CENTER
            elif col in (2, 3, 5, 6):
                c.alignment = _WRAP_C
            else:
                c.alignment = _CENTER
        ws.row_dimensions[r].height = _data_row_height(row_data)
        r += 1

    for i, text in enumerate(_USE_CODES):
        uc = ws.cell(row=r + 1 + i, column=1, value=text)
        uc.font = _NOTE_FONT; uc.alignment = _LEFT
        ws.row_dimensions[r + 1 + i].height = 16.5

    cp = ws.cell(row=r + 1 + len(_USE_CODES) + 1, column=1,
                 value="Cue Sheet Prepared By: Samraj Music Rights Management Pvt. Ltd.")
    cp.font = _FOOT_FONT; cp.alignment = _LEFT
    ws.row_dimensions[r + 1 + len(_USE_CODES) + 1].height = 14.25


def build_ascap(project, episode) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = str(episode.episode_number)[:31]
    _fill_ascap_sheet(ws, project, episode)
    return save_bytes(wb)


def build_ascap_bulk(project, episodes) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for ep in sorted(episodes, key=lambda e: e.episode_number):
        ws = wb.create_sheet(title=str(ep.episode_number)[:31])
        _fill_ascap_sheet(ws, project, ep)
    return save_bytes(wb)
