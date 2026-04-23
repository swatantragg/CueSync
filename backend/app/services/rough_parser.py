import re
from io import BytesIO
from openpyxl import load_workbook

CODE_MAP = {
    "BI": "Background Instrumental",
    "BV": "Background Vocal",
    "FV": "Featured Vocal",
    "FI": "Featured Instrumental",
    "OI": "Opening / Title",
    "VI": "Visual Instrumental",
    "VV": "Visual Vocal",
}

USAGE_TO_ENUM = {
    # IPRS codes
    "Background Instrumental": "instrumental",
    "Background Vocal":        "vocal",
    "Featured Vocal":          "vocal",
    "Featured Instrumental":   "instrumental",
    "Opening / Title":         "theme",
    "Visual Instrumental":     "visual",
    "Visual Vocal":            "visual",
    # VSG / plain-text usage labels
    "Opening Theme":           "theme",
    "Background Music":        "instrumental",
    "Visual":                  "visual",
    "Featured":                "vocal",
    "End Title":               "theme",
    "Logo":                    "theme",
}

ROLE_MAP = {"C": "Composer", "A": "Author", "E": "Publisher", "P": "Publisher"}

# VSG role normalisation
VSG_ROLE_MAP = {
    "composer":    "Composer",
    "author":      "Author",
    "publisher":   "Publisher",
    "interpreter": "Performer",
    "arranger":    "Arranger",
    "lyricist":    "Author",
}


def _s(v):
    if v is None:
        return ""
    return str(v).strip()


def _dur_to_sec(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        if 0 < v < 1:
            return int(round(v * 86400))
        return int(v)
    s = str(v).strip()
    m = re.match(r"^(\d+):(\d+):(\d+)", s)
    if m:
        return int(m[1]) * 3600 + int(m[2]) * 60 + int(m[3])
    m = re.match(r"^(\d+):(\d+)$", s)
    if m:
        return int(m[1]) * 60 + int(m[2])
    try:
        return int(float(s))
    except Exception:
        return 0


def _vsg_usage(label: str) -> str:
    l = label.strip().lower()
    if "opening" in l or "title" in l or "logo" in l or "end title" in l:
        return "theme"
    if "vocal" in l:
        return "vocal"
    if "instrumental" in l or "background music" in l or "background" in l:
        return "instrumental"
    if "visual" in l:
        return "visual"
    if "featured" in l:
        return "vocal"
    return "background"


def _ep_num_from_name(name: str) -> int | None:
    """Extract episode number from sheet name or filename."""
    m = re.search(r'ep[- _]?(\d+)', name, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r'episode[- _]?(\d+)', name, re.IGNORECASE)
    if m:
        return int(m.group(1))
    # last standalone number in name
    nums = re.findall(r'\b(\d{2,4})\b', name)
    if nums:
        return int(nums[0])
    return None


def _project_title_from_filename(filename: str) -> str:
    """'vighnaharta-ganesh-ep-301-...' → 'Vighnaharta Ganesh'"""
    name = re.sub(r'\.(xlsx?|xls)$', '', filename, flags=re.IGNORECASE)
    # strip everything from ep-NNN onwards
    name = re.split(r'ep[- _]\d+', name, flags=re.IGNORECASE)[0]
    name = re.sub(r'[-_]', ' ', name).strip()
    return name.title()


def _air_date_from_filename(filename: str) -> str | None:
    """'16th-october-2018' → '16 October 2018'"""
    m = re.search(r'(\d{1,2}(?:st|nd|rd|th))[-_ ](\w+)[-_ ](\d{4})', filename, re.IGNORECASE)
    if m:
        day  = re.sub(r'(st|nd|rd|th)$', '', m.group(1))
        mon  = m.group(2).title()
        yr   = m.group(3)
        return f"{day} {mon} {yr}"
    return None


# ── VSG format parser (one episode per sheet / file) ──────────────────────────

def _parse_vsg_sheet(rows: list, sheet_or_filename: str, hdr_idx: int) -> tuple[dict, dict]:
    """
    Parses VSG-style cue sheet.
    Header row cols: TITLE | CONTRIBUTORS | - | - | - | TIME IN | TIME OUT | DURATION | SOURCE | USAGE | LABEL

    Returns (episode_dict, project_meta_dict).
    """
    def g(ri, ci):
        r = rows[ri] if ri < len(rows) else []
        return _s(r[ci]) if len(r) > ci else ""

    # ── Project meta ─────────────────────────────────────────────────────────
    proj_title = _project_title_from_filename(sheet_or_filename)
    # Row 1 col 2 has the full title string; strip the episode part
    raw_title = g(1, 2)
    if raw_title:
        clean = re.split(r'ep[- _]?\d+', raw_title, flags=re.IGNORECASE)[0]
        clean = re.sub(r'\s+', ' ', clean).strip().rstrip('-').strip()
        if len(clean) > 3:
            proj_title = clean.title()

    network = g(3, 9) or g(3, 10) or None   # Network Station col

    proj_meta = {
        "title":        proj_title,
        "channel_name": network,
    }

    # ── Episode meta ─────────────────────────────────────────────────────────
    ep_num    = _ep_num_from_name(sheet_or_filename)
    air_date  = _air_date_from_filename(sheet_or_filename)
    total_dur = _dur_to_sec(rows[3][6]) if len(rows) > 3 and len(rows[3]) > 6 else 0

    # ── Cues ─────────────────────────────────────────────────────────────────
    cues: list[dict] = []
    current: dict | None = None
    order = 0

    for row in rows[hdr_idx + 1:]:
        if not row or not any(row):
            continue

        title_cell = _s(row[0]) if len(row) > 0 else ""
        first      = _s(row[1]) if len(row) > 1 else ""
        last       = _s(row[2]) if len(row) > 2 else ""
        role_raw   = _s(row[3]).lower() if len(row) > 3 else ""
        dur_cell   = row[7]    if len(row) > 7 else None
        usage_str  = _s(row[9]) if len(row) > 9 else ""

        if title_cell:
            # New song — first row for a song has "First name"/"Last name"/"Role" labels
            if current:
                cues.append(current)
            order += 1
            current = {
                "song_title":  title_cell,
                "usage_type":  _vsg_usage(usage_str),
                "usage_label": usage_str,
                "duration_sec": _dur_to_sec(dur_cell),
                "usage_count": 1,
                "song_code":   None,
                "order_index": order,
                "contributors": [],
            }
            # The first row cols 1-3 are label headers, not actual contributor data

        elif current and first and first.lower() not in ("first name", ""):
            # Contributor row
            name = (first + (" " + last if last else "")).strip()
            role = VSG_ROLE_MAP.get(role_raw, role_raw.title() or "Composer")
            if name:
                current["contributors"].append({
                    "name":          name,
                    "role":          role,
                    "society":       None,
                    "share_percent": 0,
                    "ipi_number":    None,
                })

    if current:
        cues.append(current)

    ep = {
        "episode_number":              ep_num,
        "title":                       None,
        "air_date":                    air_date,
        "total_duration_sec":          total_dur or None,
        "musical_duration_sec":        None,
        "bg_instrumental_duration_sec": None,
        "bg_vocal_duration_sec":       None,
        "cues":                        cues,
    }
    return ep, proj_meta


# ── IPRS / standard format helpers ────────────────────────────────────────────

def _parse_iprs_sheet(rows: list, sheet_name: str, project_meta: dict, episodes: list) -> None:
    """Parse one IPRS-style sheet. Mutates project_meta and episodes in-place."""
    ep_num = None
    try:
        ep_num = int(re.sub(r"\D", "", sheet_name) or 0) or None
    except Exception:
        ep_num = None

    ep = {
        "episode_number": ep_num,
        "title": None,
        "air_date": None,
        "total_duration_sec": None,
        "musical_duration_sec": None,
        "bg_instrumental_duration_sec": None,
        "bg_vocal_duration_sec": None,
        "cues": [],
    }
    ex: dict = {}
    song_start = -1

    def match(label, val_raw):
        L = label.upper()
        val = _s(val_raw)
        if L.startswith("TITLE"):
            m = re.search(r"EP[- ]?(\d+)", val.upper())
            if m and not ep["episode_number"]:
                ep["episode_number"] = int(m.group(1))
            m2 = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", val)
            if m2:
                ep["air_date"] = m2.group(1)
            if val:
                ep["title"] = val
        elif "DIRECTOR" in L:
            if val: ex["director"] = val
        elif "GENRE" in L:
            if val: ex["genre"] = val
        elif "LANGUAGE" in L:
            if val: ex["language"] = val
        elif "BANNER" in L or "PRODUCTION COMPANY" in L:
            if val: ex["production_company"] = val
        elif "PRINCIPAL ACTOR" in L or L.startswith("ACTOR"):
            if val: ex["actors"] = val
        elif L.startswith("PRODUCER"):
            if val: ex["producer"] = val
        elif "BACKGROUND MUSIC COMPOSER" in L:
            if val: ex["bg_music_composer"] = val
        elif "TOTAL" in L and ("MOVIE" in L or "EPISODE" in L) and "DURATION" in L:
            ep["total_duration_sec"] = _dur_to_sec(val_raw)
        elif "TOTAL MUSICAL DURATION" in L:
            ep["musical_duration_sec"] = _dur_to_sec(val_raw)
        elif "BACKGROUND INSTRUMENTAL" in L:
            ep["bg_instrumental_duration_sec"] = _dur_to_sec(val_raw)
        elif "BACKGROUND VOCAL" in L and "INSTRUMENTAL" not in L:
            ep["bg_vocal_duration_sec"] = _dur_to_sec(val_raw)

    for i, row in enumerate(rows[:40]):
        a0 = _s(row[0])
        if "SONG TITLE" in a0.upper():
            song_start = i + 1
            break
        if a0:
            match(a0, row[1] if len(row) > 1 else None)
        a6 = _s(row[6]) if len(row) > 6 else ""
        if a6:
            match(a6, row[7] if len(row) > 7 else None)

    if song_start > 0:
        current = None
        order = 0
        for row in rows[song_start:]:
            title = _s(row[0])
            upper0 = title.upper()
            if upper0.startswith("CODES") or "CHARACTERISTICS :-" in upper0 or upper0.startswith("NOTE:"):
                break
            code   = _s(row[1])
            role_c = _s(row[5])
            name   = _s(row[6])
            if title:
                if current:
                    ep["cues"].append(current)
                order += 1
                usage = CODE_MAP.get(code.upper(), code or "Background Instrumental")
                current = {
                    "song_title":  title,
                    "usage_type":  USAGE_TO_ENUM.get(usage, "background"),
                    "usage_label": usage,
                    "usage_count": int(row[2]) if isinstance(row[2], (int, float)) else 1,
                    "song_code":   _s(row[3]) or None,
                    "duration_sec": _dur_to_sec(row[4]),
                    "order_index": order,
                    "contributors": [],
                }
            if current and name and role_c:
                current["contributors"].append({
                    "name":          name,
                    "role":          ROLE_MAP.get(role_c.upper(), role_c),
                    "society":       _s(row[7]) or None,
                    "share_percent": float(row[8]) if isinstance(row[8], (int, float)) else 0,
                    "ipi_number":    _s(row[9]) or None,
                })
        if current:
            ep["cues"].append(current)

    if ep["episode_number"]:
        episodes.append(ep)

    for k, v in ex.items():
        if v and not project_meta.get(k):
            project_meta[k] = v


# ── Public entry point ────────────────────────────────────────────────────────

def parse_rough_workbook(data: bytes, filename: str = "") -> dict:
    """
    Parse rough xlsx. Supports:
      - IPRS standard format (multi-sheet, one sheet per episode, SONG TITLE header)
      - VSG format         (single-sheet, TITLE / CONTRIBUTORS header)

    Returns {'meta': {...}, 'episodes': [{...}]}.
    """
    wb = load_workbook(BytesIO(data), data_only=True)
    project_meta: dict = {}
    episodes: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r) + [None] * max(0, 12 - len(r)))

        # ── Detect VSG format ─────────────────────────────────────────────────
        # Header row: col0="TITLE", col1="CONTRIBUTORS"
        vsg_hdr = next(
            (i for i, r in enumerate(rows[:15])
             if _s(r[0]).upper() == "TITLE" and _s(r[1]).upper() == "CONTRIBUTORS"),
            None,
        )

        if vsg_hdr is not None:
            # Use filename for episode-number / air-date extraction when available
            hint = filename or sheet_name
            ep, pm = _parse_vsg_sheet(rows, hint, vsg_hdr)
            if ep.get("episode_number"):
                episodes.append(ep)
            for k, v in pm.items():
                if v and not project_meta.get(k):
                    project_meta[k] = v
        else:
            # ── Standard IPRS format ──────────────────────────────────────────
            _parse_iprs_sheet(rows, sheet_name, project_meta, episodes)

    return {"meta": project_meta, "episodes": episodes}
